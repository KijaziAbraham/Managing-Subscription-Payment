from .models import SoftwareCategory, SoftwareEdition, SoftwareVersion, Software, CompanyUser, AuditLog, UserCategory
from .forms import SoftwareCategoryForm, SoftwareEditionForm, SoftwareVersionForm, SoftwareForm, CustomPasswordChangeForm, CompanyUserForm, ImportForm, SoftwareImportForm, AddonForm
from django.contrib.auth import authenticate, login as auth_login,logout as auth_logout
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import update_session_auth_hash
from django.db.models import BooleanField, Case, Value, When
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.forms import UserCreationForm
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter
from django.http import HttpResponse
from reportlab.lib.units import inch
from django.contrib import messages
from reportlab.pdfgen import canvas
from django.utils import timezone
from reportlab.lib import colors
from django.urls import reverse
from datetime import timedelta
from django.http import JsonResponse
from django.contrib.auth.models import User
from django.contrib.auth.decorators import user_passes_test
from django.shortcuts import render
from django.db.models import Q
from django.contrib.auth.decorators import permission_required
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from dateutil.relativedelta import relativedelta  
from .models import Addon
from django.template.loader import render_to_string
from django.db import transaction
from django.utils.timezone import now as timezone_now
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer, Image
from reportlab.lib.units import inch
from django.contrib.staticfiles.storage import staticfiles_storage
from django.http import HttpResponse
from django.utils import timezone
from datetime import timedelta
from DuxteSubscriptions.models import CompanyUser, UserCategory
from openpyxl.styles import Font, Alignment
from openpyxl.comments import Comment
from datetime import datetime
from django.contrib import messages
from django.db import transaction
import openpyxl
import pandas as pd
import io
import re




def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)

            if user is not None:
                auth_login(request, user)

                # Check for existing company users and software
                if CompanyUser.objects.exists():
                    return redirect('dashboard')
                elif Software.objects.exists():
                    return redirect('create_company_user')
                else:
                    return redirect('add_software')
            else:
                messages.error(request, 'Invalid username or password.')
        else:
            messages.error(request, 'Please correct the error, Invalid username or password.')
    else:
        form = AuthenticationForm()

    context = {
        'form': form
    }

    return render(request, 'DuxteSubscriptions/login.html', context)

def logout_view(request):
    auth_logout(request)
    return redirect('login')

@login_required
def create_company_user(request):
    if request.user.is_superuser:
        user_categories = Software.objects.values_list('category', flat=True)
    else:
        # Get the categories assigned to the logged-in user
        user_categories = UserCategory.objects.filter(user=request.user).values_list('category', flat=True)

    if request.method == 'POST':
        form = CompanyUserForm(request.POST)
        if form.is_valid():
            company_user = form.save(commit=False)
            company_user.is_active = True

            # Ensure selected software is within the user's assigned categories
            software = form.cleaned_data.get('software')
            if not request.user.is_superuser and software.category_id not in user_categories:
                messages.error(request, "You can only assign software from categories you have access to.")
                return render(request, 'DuxteSubscriptions/create_company.html', {'form': form})

            company_user.save()

            # Handle software_addons field - using set() for many-to-many relationship
            selected_addons = form.cleaned_data.get('software_addons')
            if selected_addons:
                company_user.software_addons.set(selected_addons)
            else:
                company_user.software_addons.clear()

            company_user.save()

            AuditLog.objects.create(
                action='CREATE',
                user=request.user,
                company_user=company_user,
                details=(
                    f"User '{company_user.customer_name}' was successfully created on "
                    f"{timezone.now().strftime('%Y-%m-%d %H:%M:%S')} by {request.user.get_full_name()} "
                    f"(Username: {request.user.username})."
                )
            )
            messages.success(request, f'The company user "{company_user.customer_name}" was successfully created.')
            return redirect('company_user_list')
        else:
            software = form.cleaned_data.get('software')
            if software:
                form.fields['software_edition'].queryset = SoftwareEdition.objects.filter(software=software)
                form.fields['software_version'].queryset = SoftwareVersion.objects.filter(software=software)
                form.fields['software_addons'].queryset = software.addons.all()  # Use the reverse relationship

    else:
        form = CompanyUserForm()

    form.fields['software'].queryset = Software.objects.filter(category__in=user_categories)

    return render(request, 'DuxteSubscriptions/create_company.html', {'form': form})

def load_editions_and_versions(request):
    software_id = request.GET.get('software') 

    if software_id:
        try:
            software = Software.objects.get(id=software_id)

            editions = software.editions.all().values('id', 'name')
            versions = software.versions.all().values('id', 'name')
            addons = software.addons.all().values('id', 'name')

            return JsonResponse({
                'editions': list(editions),
                'versions': list(versions),
                'addons': list(addons),
            })
        except Software.DoesNotExist:
            return JsonResponse({'error': 'Software not found'}, status=404)

    return JsonResponse({'error': 'Invalid request'}, status=400)

def update_company_user(request, pk):
    company_user = get_object_or_404(CompanyUser, pk=pk)

    if request.user.is_superuser:
        user_categories = Software.objects.values_list('category', flat=True)
    else:
        user_categories = UserCategory.objects.filter(user=request.user).values_list('category', flat=True)

    if request.method == 'POST':
        form = CompanyUserForm(request.POST, instance=company_user)
        
        if form.is_valid():
            updated_user = form.save(commit=False)
            today = timezone.now().date()
            
            if updated_user.end_of_subscription != company_user.end_of_subscription:
                # Reset reminder status if the subscription has been updated
                updated_user.last_reminder_sent = None  
                updated_user.last_renewal_date = today  

            updated_user.is_active = True
            updated_user.save()

            selected_addons = form.cleaned_data.get('software_addons')
            if selected_addons:
                updated_user.software_addons.set(selected_addons)
            else:
                updated_user.software_addons.clear()

            AuditLog.objects.create(
                action='UPDATE',
                user=request.user,
                company_user=updated_user,
                details=(f"User '{company_user.customer_name}' (ID: {company_user.id}) was updated on "
                         f"{timezone.now().strftime('%Y-%m-%d %H:%M:%S')} by {request.user.get_full_name()} "
                         f"(Username: {request.user.username}).")
            )
            
            return redirect('company_user_detail', pk=updated_user.pk)
    else:
        form = CompanyUserForm(instance=company_user)

        form.fields['software_edition'].initial = company_user.software_edition
        form.fields['software_version'].initial = company_user.software_version
        form.fields['software_addons'].initial = company_user.software_addons.all()

    form.fields['software'].queryset = Software.objects.filter(category__in=user_categories)

    return render(request, 'DuxteSubscriptions/update_company_user.html', {'form': form, 'company_user': company_user})

@login_required
def company_user_list(request, category=None):
    if request.user.is_superuser:
        company_users = CompanyUser.objects.filter(deleted_at__isnull=True)
    else:
        user_categories = UserCategory.objects.filter(user=request.user).values_list('category', flat=True)

        if category:
            if category in user_categories:
                company_users = CompanyUser.objects.filter(software__category=category, deleted_at__isnull=True)
            else:
                messages.error(request, "You do not have access to this category.")
                return redirect('company_user_list')
        else:
            company_users = CompanyUser.objects.filter(software__category__in=user_categories, deleted_at__isnull=True)

    if not company_users.exists():
        return redirect('create_company_user')

    return render(request, 'DuxteSubscriptions/company_user_list.html', {
        'company_users': company_users,
        'category': category
    })

@login_required
def company_user_list_with_category(request, category):
    today = timezone.now().date()
    two_months_from_now = today + timedelta(days=60)


    if request.user.is_superuser:
        base_query = CompanyUser.objects.filter(deleted_at__isnull=True)
    else:
        user_categories = UserCategory.objects.filter(user=request.user).values_list('category', flat=True)

        base_query = CompanyUser.objects.filter(software__category__in=user_categories, deleted_at__isnull=True)

    if category == 'follow_up':
        company_users = base_query.filter(
            end_of_subscription__gt=today,
            end_of_subscription__lte=two_months_from_now
        )
    elif category == 'active':
        company_users = base_query.filter(is_active=True)
    elif category == 'valid':
        company_users = base_query.filter(end_of_subscription__gt=today)
    elif category == 'expired':
        company_users = base_query.filter(end_of_subscription__lte=today)
    else:
        company_users = base_query.none()

    # Normalize category name for display (e.g., 'follow_up' becomes 'Follow Up')
    display_category = category.replace('_', ' ').title()

    context = {
        'company_users': company_users,
        'category': display_category  
    }

    return render(request, 'DuxteSubscriptions/company_user_list.html', context)

def superuser_required(user):
    return user.is_superuser

@user_passes_test(superuser_required)
def soft_delete_user(request, user_id):
    user = get_object_or_404(CompanyUser, id=user_id)
    user.deleted_at = timezone.now()
    user.save()
    
    AuditLog.objects.create(
        action='DELETE',
        user=request.user,
        company_user=user,
        details=(
            f"User '{user.customer_name}' (ID: {user.id}) was soft deleted on "
            f"{timezone.now().strftime('%Y-%m-%d %H:%M:%S')} by {request.user.get_full_name()} "
            f"(Username: {request.user.username})."
        )
    )
    
    messages.success(request, f'The company user "{user.customer_name}" was successfully soft deleted.')
    return redirect('company_user_list')

@user_passes_test(superuser_required)
def soft_delete_software(request, pk):
    software = get_object_or_404(Software, pk=pk)
    software.is_deleted = True  
    software.save()
    return redirect('software_list') 

@user_passes_test(superuser_required)
def restore_user(request, user_id):
    user = get_object_or_404(CompanyUser, id=user_id)
    if user.deleted_at:
        user.deleted_at = None
        user.save()
        
        AuditLog.objects.create(
            action='RESTORE',
            user=request.user,
            company_user=user,
            details=(
                f"User '{user.customer_name}' (ID: {user.id}) was restored on "
                f"{timezone.now().strftime('%Y-%m-%d %H:%M:%S')} by {request.user.get_full_name()} "
                f"(Username: {request.user.username})."
            )
        )
        
        messages.success(request, f'The company user "{user.customer_name}" was successfully restored.')
    return redirect('deleted_user_list')

def deleted_user_list(request):
    deleted_users = CompanyUser.objects.filter(deleted_at__isnull=False)
    return render(request, 'DuxteSubscriptions/deleted_user_list.html', {
        'deleted_users': deleted_users
    })

def company_user_detail(request, pk):
    company_user = get_object_or_404(CompanyUser, pk=pk)
    return render(request, 'DuxteSubscriptions/company_user_detail.html', {
        'company_user': company_user,
    })

@user_passes_test(superuser_required)
def add_software_category(request):
    if request.method == 'POST':
        form = SoftwareCategoryForm(request.POST)
        if form.is_valid():
            categories = form.cleaned_data['name'].split(',')
            existing_categories = {category.name for category in SoftwareCategory.objects.all()}
            new_categories = []

            for category_name in categories:
                category_name = category_name.strip()
                if category_name:
                    if category_name not in existing_categories:
                        SoftwareCategory.objects.create(name=category_name)
                        new_categories.append(category_name)
                    else:
                        messages.warning(request, f'The category "{category_name}" already exists. It will not be added.')

            if new_categories:
                messages.success(request, f'The categories "{", ".join(new_categories)}" were successfully added.')
            return redirect('add_software_category')
    else:
        form = SoftwareCategoryForm()

    categories = SoftwareCategory.objects.all()

    return render(request, 'DuxteSubscriptions/add_software_category.html', {'form': form, 'categories': categories})

@user_passes_test(superuser_required)
def add_software_edition(request):
    if request.method == 'POST':
        form = SoftwareEditionForm(request.POST)
        if form.is_valid():
            editions = form.cleaned_data['name'].split(',')
            existing_editions = {edition.name for edition in SoftwareEdition.objects.all()}
            new_editions = []

            for edition_name in editions:
                edition_name = edition_name.strip()
                if edition_name:
                    if edition_name not in existing_editions:
                        SoftwareEdition.objects.create(name=edition_name)
                        new_editions.append(edition_name)

            if new_editions:
                messages.success(request, f'The editions "{", ".join(new_editions)}" were successfully added.')
            return redirect('add_software_edition')
    else:
        form = SoftwareEditionForm()

    editions = SoftwareEdition.objects.all()

    return render(request, 'DuxteSubscriptions/add_software_edition.html', {'form': form, 'editions': editions})

@user_passes_test(superuser_required)
def add_software_version(request):
    if request.method == 'POST':
        form = SoftwareVersionForm(request.POST)
        if form.is_valid():
            versions = form.cleaned_data['name'].split(',')
            existing_versions = {version.name for version in SoftwareVersion.objects.all()}
            new_versions = []

            for version_name in versions:
                version_name = version_name.strip()
                if version_name:
                    if version_name not in existing_versions:
                        SoftwareVersion.objects.create(name=version_name)
                        new_versions.append(version_name)

            if new_versions:
                messages.success(request, f'The versions "{", ".join(new_versions)}" were successfully added.')
            return redirect('add_software_version')
    else:
        form = SoftwareVersionForm()

    versions = SoftwareVersion.objects.all()

    return render(request, 'DuxteSubscriptions/add_software_version.html', {'form': form, 'versions': versions})

@user_passes_test(superuser_required)
def create_addon(request):
    if request.method == 'POST':
        form = AddonForm(request.POST)
        if form.is_valid():
            addons = form.cleaned_data['name'].split(',')
            existing_addons = {addon.name for addon in Addon.objects.all()}
            new_addons = []

            for addon_name in addons:
                addon_name = addon_name.strip()
                if addon_name:
                    if addon_name not in existing_addons:
                        Addon.objects.create(name=addon_name)
                        new_addons.append(addon_name)
                    else:
                        messages.warning(request, f'The addon "{addon_name}" already exists. It will not be added.')

            if new_addons:
                messages.success(request, f'The addons "{", ".join(new_addons)}" were successfully added.')
            return redirect('create_addon')
    else:
        form = AddonForm()

    addons = Addon.objects.all()

    return render(request, 'DuxteSubscriptions/add_addon.html', {'form': form, 'addons': addons})

@require_POST
def delete_software_category(request, category_id):
    try:
        category = SoftwareCategory.objects.get(id=category_id)
        category.delete()
        messages.success(request, f'Category "{category.name}" deleted successfully.')
    except SoftwareCategory.DoesNotExist:
        messages.error(request, 'Category not found.')
    return JsonResponse({'success': True})

@require_POST
def delete_software_edition(request, edition_id):
    try:
        edition = SoftwareEdition.objects.get(id=edition_id)
        edition.delete()
        messages.success(request, f'Edition "{edition.name}" deleted successfully.')
    except SoftwareEdition.DoesNotExist:
        messages.error(request, 'Edition not found.')
    return JsonResponse({'success': True})

@require_POST
def delete_software_version(request, version_id):
    try:
        version = SoftwareVersion.objects.get(id=version_id)
        version.delete()
        messages.success(request, f'Version "{version.name}" deleted successfully.')
    except SoftwareVersion.DoesNotExist:
        messages.error(request, 'Version not found.')
    return JsonResponse({'success': True})

@require_POST
def delete_addon(request, addon_id):
    try:
        addon = Addon.objects.get(id=addon_id)
        addon.delete()
        messages.success(request, f'Addon "{addon.name}" deleted successfully.')
    except Addon.DoesNotExist:
        messages.error(request, 'Addon not found.')
    return JsonResponse({'success': True})

def edit_software_category(request, category_id):
    if request.method == 'POST':
        category = get_object_or_404(SoftwareCategory, id=category_id)
        name = request.POST.get('name')
        category.name = name
        category.save()
        return JsonResponse({'message': 'Category updated successfully.'})

def edit_software_edition(request, edition_id):
    if request.method == 'POST':
        edition = get_object_or_404(SoftwareEdition, id=edition_id)
        name = request.POST.get('name')
        edition.name = name
        edition.save()
        return JsonResponse({'message': 'Edition updated successfully.'})

def edit_software_version(request, version_id):
    if request.method == 'POST':
        version = get_object_or_404(SoftwareVersion, id=version_id)
        name = request.POST.get('name')
        version.name = name
        version.save()
        return JsonResponse({'message': 'Version updated successfully.'})
    
def edit_software_addon(request, addon_id):
    if request.method == 'POST':
        addon = get_object_or_404(Addon, id=addon_id)
        name = request.POST.get('name')
        addon.name = name
        addon.save()
        return JsonResponse({'message': 'Addon updated successfully.'})

@user_passes_test(superuser_required)
def add_software(request):
    available_editions = SoftwareEdition.objects.all()
    available_versions = SoftwareVersion.objects.all()
    available_addons = Addon.objects.all()

    if request.method == 'POST':
        form = SoftwareForm(request.POST)

        if form.is_valid():
            software_name = form.cleaned_data.get('name').strip()

            # Check if a software with the same name already exists
            if Software.objects.filter(name__iexact=software_name).exists():
                messages.error(request, f'Software with the name "{software_name}" already exists.')
                return render(request, 'DuxteSubscriptions/add_software.html', {'form': form, 
                    'available_editions': available_editions,
                    'available_versions': available_versions,
                    'available_addons': available_addons})

            software = form.save()
            
            editions = request.POST.getlist('editions')
            versions = request.POST.getlist('versions')
            addons = request.POST.getlist('addons')

            # Save the ManyToMany relationships
            software.editions.set(editions)
            software.versions.set(versions)
            software.addons.set(addons)

            AuditLog.objects.create(
                action='CREATE',
                user=request.user,
                details=(f"Software '{software.name}' was successfully created on "
                         f"{timezone.now().strftime('%Y-%m-%d %H:%M:%S')} by {request.user.get_full_name()} "
                         f"(Username: {request.user.username}).")
            )
            messages.success(request, f'The software "{software.name}" was successfully added.')
            return redirect('software_list')
    else:
        form = SoftwareForm()

    return render(request, 'DuxteSubscriptions/add_software.html', {
        'form': form,
        'available_editions': available_editions,
        'available_versions': available_versions,
        'available_addons': available_addons
    })

@user_passes_test(superuser_required)
def software_list(request):
    software_list = Software.objects.prefetch_related('addons', 'versions', 'editions').all()
    software_list = Software.objects.filter(is_deleted=False)
    return render(request, 'DuxteSubscriptions/software_list.html', {'software_list': software_list})

def edit_software(request, pk):
    software = get_object_or_404(Software, pk=pk)


    available_editions = SoftwareEdition.objects.all()
    available_versions = SoftwareVersion.objects.all()
    available_addons = Addon.objects.all()

    if request.method == 'POST':
        form = SoftwareForm(request.POST, instance=software)
        if form.is_valid():
            form.save()

            software.editions.set(request.POST.getlist('editions'))
            software.versions.set(request.POST.getlist('versions'))
            software.addons.set(request.POST.getlist('addons'))
            return redirect('software_list')
    else:
        form = SoftwareForm(instance=software)

    return render(request, 'DuxteSubscriptions/edit_software.html', {
        'form': form,
        'available_editions': available_editions,
        'available_versions': available_versions,
        'available_addons': available_addons,
    })

@login_required
def export_to_pdf(request, category=None):
    if request.user.is_superuser:
        user_categories = None
    else:
        user_categories = UserCategory.objects.filter(user=request.user).values_list('category', flat=True)

    if category is None:
        category = 'all'
    category = category.lower().replace(' ', '_')

    filename = f"company user list {category}.pdf"
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    doc = SimpleDocTemplate(response, pagesize=letter)
    elements = []

    styles = getSampleStyleSheet()

    # different sections 
    header_style = ParagraphStyle('Header', fontName='Helvetica-Bold', fontSize=7, leading=9, alignment=0, spaceAfter=5)
    normal_style = ParagraphStyle('BodyText', parent=styles['BodyText'], fontName='Helvetica', fontSize=7, leading=9)
    title_style = ParagraphStyle('Title', parent=styles['Title'], fontName='Helvetica-Bold', fontSize=10, leading=12, alignment=1, spaceAfter=10)

    line_top = Table([['']], colWidths=[7.5 * inch])
    line_top.setStyle(TableStyle([('LINEABOVE', (0, 0), (-1, 0), 1, colors.black)]))
    elements.append(line_top)

    logo_path = 'static/img/favicon2.png'
    logo = Image(logo_path, width=1 * inch, height=1 * inch)

    company_info = [
    "Duxte Limited",
    "Plot 7B, Avocado Street",
    "Mikocheni",
    "P.O.Box 35984, Dar es Salaam",
    "Tanzania"
    ]
    company_info_paragraph = Paragraph("<br />".join(company_info), normal_style)

    contact_info = [
        "Phone:  +255 745 000 555",
        "Fax:  +255 745 000 555",
        "Email: biz@duxte.com",
        "Website: www.duxte.com"
    ]
    contact_info_paragraph = Paragraph("<br />".join(contact_info), normal_style)

    # Create a table for the header with logo, company address, and contact info
    header_table_data = [[logo, company_info_paragraph, contact_info_paragraph]]
    header_table = Table(header_table_data, colWidths=[1 * inch, 3 * inch, 3.5 * inch])  
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),  
        ('ALIGN', (1, 0), (1, 0), 'LEFT'),  
        ('ALIGN', (2, 0), (2, 0), 'RIGHT'), 
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(header_table)

    elements.append(Spacer(1, 5))

    title_text = "COMPANY USER LIST"
    title = Paragraph(title_text, title_style)
    elements.append(title)

    # space before the table
    elements.append(Spacer(1, 6))

    headers = ["#", "Company", "Email", "Start Date", "End Date", "Status"]
    data = [headers]

    today = timezone.now().date()
    two_months_from_now = today + timedelta(days=60)

    if category == 'follow_up':
        company_users = CompanyUser.objects.filter(end_of_subscription__gt=today, end_of_subscription__lte=two_months_from_now)
    elif category == 'active':
        company_users = CompanyUser.objects.filter(is_active=True)
    elif category == 'valid':
        company_users = CompanyUser.objects.filter(end_of_subscription__gte=timezone.now())
    elif category == 'expired':
        company_users = CompanyUser.objects.filter(end_of_subscription__lte=timezone.now())
    else:
        company_users = CompanyUser.objects.all()

    if not request.user.is_superuser:
        company_users = company_users.filter(software__category__in=user_categories)

    # Populate table 
    for index, company_user in enumerate(company_users, start=1):
        row = [
            str(index),
            company_user.customer_name,
            company_user.email1,
            company_user.date_of_subscription.strftime('%Y-%m-%d') if company_user.date_of_subscription else '',
            company_user.end_of_subscription.strftime('%Y-%m-%d') if company_user.end_of_subscription else '',
            'Valid' if company_user.subscription_valid else 'Expired'
        ]
        data.append(row)

    table = Table(data, hAlign='CENTER', colWidths=[0.5 * inch, 3.5 * inch, 1.8 * inch, 0.8 * inch, 0.8 * inch, 0.6 * inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 0), (-1, -1), 7),  
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('WORDWRAP', (0, 1), (-1, -1), True),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 6))

    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontName='Helvetica', fontSize=7, alignment=1, spaceAfter=6)
    footer_text = "Contact us at:  biz@duxte.com| © 2024 DUXTE LTD"
    footer = Paragraph(footer_text, footer_style)
    elements.append(Spacer(1, 6))
    elements.append(footer)

    doc.build(elements)
    return response

@login_required
def export_to_excel(request, category=None):
    if request.user.is_superuser:
        user_categories = None  
    else:
        user_categories = UserCategory.objects.filter(user=request.user).values_list('category', flat=True)

    if category is None:
        category = 'all'
    category = category.lower().replace(' ', '_')

    filename = f"company_user_list_{category}.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Company Users'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
    alignment = Alignment(horizontal='center', vertical='center')

    headers = [
        "Customer Name", "Customer Account", "Base Serial Number", "Contact Person", "Phone Number", 
        "Email 1", "Email 2", "Date of Registration", "Date of Subscription", 
        "Subscription Duration", "Is Active", "Software Name", 
        "Software Edition", "Software Version", "Software Add-ons","Subscription End", "Status"
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    column_widths = [20, 20, 20, 20, 15, 25, 25, 25, 25, 30, 15, 20, 20, 20, 40,25,15]
    for i, column_width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = column_width

    today = timezone.now().date()
    two_months_from_now = today + timedelta(days=60)

    if category == 'follow_up':
        company_users = CompanyUser.objects.filter(end_of_subscription__gt=today, end_of_subscription__lte=two_months_from_now)
    elif category == 'active':
        company_users = CompanyUser.objects.filter(is_active=True)
    elif category == 'valid':
        company_users = CompanyUser.objects.filter(end_of_subscription__gte=timezone.now())
    elif category == 'expired':
        company_users = CompanyUser.objects.filter(end_of_subscription__lte=timezone.now())
    else:
        company_users = CompanyUser.objects.all()

    if not request.user.is_superuser:
        company_users = company_users.filter(software__category__in=user_categories)

    try:
        for index, company_user in enumerate(company_users, start=1):
            addons = ', '.join([addon.name for addon in company_user.software_addons.all()])

            row = [
                company_user.customer_name,
                company_user.customer_account or '',
                company_user.base_serial_number or '',
                company_user.contact,
                company_user.phone_number,
                company_user.email1,
                company_user.email2 or '',
                company_user.date_of_registration.strftime('%Y-%m-%d') if company_user.date_of_registration else '',
                company_user.date_of_subscription.strftime('%Y-%m-%d') if company_user.date_of_subscription else '',
                company_user.subscription_duration,
                "True" if company_user.is_active else "False",
                company_user.software.name if company_user.software else '',
                company_user.software_edition.name if company_user.software_edition else '',
                company_user.software_version.name if company_user.software_version else '',
                addons,
                company_user.end_of_subscription.strftime('%Y-%m-%d') if company_user.end_of_subscription else '',
                "Valid" if company_user.subscription_valid else "Expired"

            ]
            ws.append(row)

            for cell in ws[index + 1]:
                cell.alignment = alignment

    except Exception as e:
        ws.append(['Error:', str(e)])

    wb.save(response)
    return response

def toggle_status(request, pk):
    company_user = get_object_or_404(CompanyUser, pk=pk)
    company_user.is_active = not company_user.is_active
    company_user.save()
    return redirect('company_user_list')

@login_required
def dashboard(request):
    today = timezone.now().date()
    two_months_from_now = today + timedelta(days=60)

    if request.user.is_superuser:
        all_users = CompanyUser.objects.filter(deleted_at__isnull=True)
    else:
        user_categories = UserCategory.objects.filter(user=request.user).values_list('category', flat=True)

        all_users = CompanyUser.objects.filter(
            deleted_at__isnull=True,
            software__category__in=user_categories
        )

    follow_up_users = all_users.filter(
        end_of_subscription__gt=today,
        end_of_subscription__lte=two_months_from_now
    )
    follow_up_count = follow_up_users.count()

    reminders = [user for user in all_users if user.should_send_reminder()]
    reminder_count = len(reminders)

    active_count = all_users.filter(is_active=True).count()

    valid_count = all_users.filter(end_of_subscription__gt=today).count()

    expired_count = all_users.filter(end_of_subscription__lte=today).count()

    suspended_count = all_users.filter(is_active=False).count()

    context = {
        'follow_up_count': follow_up_count,
        'follow_up_users': follow_up_users,
        'reminder_count': reminder_count,
        'reminders': reminders,
        'active_count': active_count,
        'valid_count': valid_count,
        'expired_count': expired_count,
        'suspended_count': suspended_count,
    }

    return render(request, 'DuxteSubscriptions/dashboard.html', context)

def follow_up_customers(request):
    today = timezone.now().date()
    two_months_from_now = today + timedelta(days=60)
    customers = CompanyUser.objects.filter(end_of_subscription__gt=today, end_of_subscription__lte=two_months_from_now, deleted_at__isnull=True)
    return render(request, 'DuxteSubscriptions/company_user_list.html', {'company_users': customers})

def active_customers(request):
    customers = CompanyUser.objects.filter(is_active=True, deleted_at__isnull=True)
    return render(request, 'DuxteSubscriptions/company_user_list.html', {'company_users': customers})

def valid_customers(request):
    today = timezone.now().date()
    customers = CompanyUser.objects.filter(end_of_subscription__gt=today, deleted_at__isnull=True)
    return render(request, 'DuxteSubscriptions/company_user_list.html', {'company_users': customers})

def expired_customers(request):
    today = timezone.now().date()
    customers = CompanyUser.objects.filter(end_of_subscription__lte=today, deleted_at__isnull=True)
    return render(request, 'DuxteSubscriptions/company_user_list.html', {'company_users': customers})

def suspended_customers(request):
    customers = CompanyUser.objects.filter(Q(is_active=False) | Q(is_deleted=True))
    return render(request, 'DuxteSubscriptions/company_user_list.html', {'company_users': customers})

@login_required
def change_password(request):
    if request.method == 'POST':
        form = CustomPasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Your password has been successfully updated!')
            return redirect('login')
        else:
            messages.error(request, 'Please correct the error below.')
    else:
        form = CustomPasswordChangeForm(user=request.user)
    return render(request, 'DuxteSubscriptions/change_password.html', {'form': form})

@login_required
def software_trends(request):
    if request.user.is_superuser:
        software_data = Software.objects.all()
    else:
        user_categories = UserCategory.objects.filter(user=request.user).values_list('category', flat=True)
        
        software_data = Software.objects.filter(category__in=user_categories)

    labels = []
    values = []

    for software in software_data:
        labels.append(software.name)  
        count = CompanyUser.objects.filter(software=software).count()
        values.append(count) 

    data = {
        'labels': labels,
        'values': values
    }
    return JsonResponse(data)

@user_passes_test(lambda u: u.is_superuser)
def audit_log_list(request):
    audit_logs = AuditLog.objects.all().order_by('-timestamp')
    return render(request, 'DuxteSubscriptions/audit_log_list.html', {'audit_logs': audit_logs})

def import_template_view(request):
    return render(request, 'DuxteSubscriptions/import_template.html')

def validate_phone_number(phone_number):
    digits_only = re.sub(r'\D', '', phone_number)
    return 10 <= len(digits_only) <= 16

def parse_date(date_str):
    date_formats = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"]
    for fmt in date_formats:
        try:
            return datetime.strptime(date_str.split(' ')[0], fmt).date() 
        except ValueError:
            continue
    return None

def download_company_template(request):
    columns = [
        'Customer Name', 'Customer Account', 'Base Serial Number', 'Contact Person', 'Phone Number', 'Email 1', 'Email 2',
        'Date of Registration (YYYY-MM-DD)', 'Date of Subscription (YYYY-MM-DD)', 'Subscription Duration (Months)',
        'Is Active (True/False)', 'Software Name', 'Software Edition', 'Software Version', 'Software Add-ons (Comma Separated)'
    ]
    
    df = pd.DataFrame(columns=columns)

    output = io.BytesIO()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Company Template"

    for row in dataframe_to_rows(df, index=False, header=True):
        worksheet.append(row)

    column_widths = [20, 20, 20, 20, 15, 25, 25, 25, 25, 30, 15, 20, 20, 20, 40]
    for i, column_width in enumerate(column_widths, 1):
        worksheet.column_dimensions[chr(64 + i)].width = column_width

    header_font = Font(bold=True, color='FFFFFF')  
    header_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid') 
    alignment = Alignment(horizontal='center', vertical='center') 

    headers = worksheet[1]  
    for cell in headers:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment

    workbook.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=company_import_template.xlsx'
    return response

def import_users(request):
    errors = []
    if request.method == 'POST':
        form = ImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            try:
                df = pd.read_excel(file, dtype=str)  
                df = df.fillna('')  
            except Exception as e:
                messages.error(request, f"Error reading Excel file: {str(e)}")
                return redirect('import_users')




        # Step 1: Validating rows
            for index, row in df.iterrows():
                row_errors = []
                software, edition, version, row_errors = validate_software_row(row, row_errors)

                subscription_duration = int(row[9]) if row[9].isdigit() else None
                if subscription_duration is None or not (1 <= subscription_duration <= 12):
                    row_errors.append("Subscription duration must be between 1 and 12 months.")

                phone_number = row[4]
                if not validate_phone_number(phone_number):
                    row_errors.append("Phone number must contain between 10 and 16 digits.")

                date_of_registration = parse_date(row[7])
                date_of_subscription = parse_date(row[8])
                if not date_of_registration:
                    row_errors.append(f"Invalid date of registration: '{row[7]}'")
                if not date_of_subscription:
                    row_errors.append(f"Invalid date of subscription: '{row[8]}'")

                if date_of_registration and date_of_subscription and date_of_registration > date_of_subscription:
                    row_errors.append("Date of registration cannot be after the date of subscription.")

                if row_errors:
                    errors.append(f"Row {index + 2}: {', '.join(row_errors)}")

          # Step 2: Saving data if no errors
            if not errors:
                try:
                    with transaction.atomic():
                        for index, row in df.iterrows():
                            try:
                                software, edition, version, _ = validate_software_row(row, [])
                                addons = Addon.objects.filter(name__in=[addon.strip() for addon in row[14].split(",") if addon.strip()])
                                company_user = CompanyUser(
                                    customer_name=row[0],
                                    customer_account=row[1],
                                    base_serial_number=row[2],
                                    contact=row[3],
                                    phone_number=row[4],
                                    email1=row[5],
                                    email2=row[6] if row[6] else None,
                                    date_of_registration=parse_date(row[7]),
                                    date_of_subscription=parse_date(row[8]),
                                    subscription_duration=subscription_duration,
                                    is_active=row[10],
                                    software=software,
                                    software_edition=edition,
                                    software_version=version,
                                )
                                company_user.save()
                                company_user.software_addons.set(addons)

                            except Exception as e:
                                errors.append(f"Error importing row {index + 2}: {str(e)}")

                    if errors:
                        log_import_failure(request.user, errors)
                        messages.error(request, 'Import failed. Check the error report.')
                        return render(request, 'DuxteSubscriptions/import_company.html', {'form': form, 'errors': errors})

                    log_import_success(request.user, len(df) - len(errors))
                    messages.success(request, 'User import process completed successfully.')
                    return redirect('company_user_list')

                except Exception as e:
                    messages.error(request, f"Import failed: {str(e)}")
                    log_import_failure(request.user, str(e))
                    return render(request, 'DuxteSubscriptions/import_company.html', {'form': form, 'errors': errors})

            else:
                messages.error(request, 'Import failed with errors. Please check the detailed error report.')
                return render(request, 'DuxteSubscriptions/import_company.html', {'form': form, 'errors': errors})

    else:
        form = ImportForm()

    return render(request, 'DuxteSubscriptions/import_company.html', {'form': form, 'errors': errors})

# Helper function to fetch and validate software data and Importation
def validate_software_row(row, row_errors):
    try:
        software = Software.objects.get(name__iexact=row[11].strip().lower())
    except Software.DoesNotExist:
        row_errors.append(f"Software '{row[11]}' does not exist.")
        software = None

    try:
        edition = SoftwareEdition.objects.get(name__iexact=row[12].strip().lower())
    except SoftwareEdition.DoesNotExist:
        row_errors.append(f"Edition '{row[12]}' does not exist.")
        edition = None

    try:
        version = SoftwareVersion.objects.get(name__iexact=row[13].strip().lower())
    except SoftwareVersion.DoesNotExist:
        row_errors.append(f"Version '{row[13]}' does not exist.")
        version = None

    return software, edition, version, row_errors

def log_import_success(user, imported_count):
    AuditLog.objects.create(
        action='IMPORT SUCCEEDED',
        user=user,
        details=f"Imported {imported_count} users on {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}."
    )

def log_import_failure(user, error_details):
    AuditLog.objects.create(
        action='IMPORT FAILED',
        user=user,
        details=f"Import failed on {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}: {error_details}."
    )

def download_software_template(request):
    columns = ['name', 'category', 'editions (comma separated)', 'versions (comma separated)', 'addons (comma separated)']
    
    df = pd.DataFrame(columns=columns)

    output = io.BytesIO()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Software Template"

    for row in dataframe_to_rows(df, index=False, header=True):
        worksheet.append(row)

    column_widths = [30, 20, 40, 40, 40]  
    for i, column_width in enumerate(column_widths, 1):
        worksheet.column_dimensions[chr(64 + i)].width = column_width

    headers = worksheet[1]
    for header in headers:
        header.font = Font(bold=True)
        header.alignment = Alignment(horizontal='center', vertical='center')

    workbook.save(output)
    output.seek(0)

    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=software_import_template.xlsx'
    return response

def import_software(request):
    if request.method == 'POST':
        form = SoftwareImportForm(request.POST, request.FILES)
        if form.is_valid():
            file = request.FILES['file']
            try:
                df = pd.read_excel(file, dtype=str) 
                df = df.fillna('')  
            except Exception as e:
                messages.error(request, f"Error reading Excel file: {str(e)}")
                AuditLog.objects.create(
                    user=request.user,
                    action="IMPORT FAILED",
                    details=f"Error reading Excel file: {str(e)}"
                )
                return render(request, 'DuxteSubscriptions/import_software.html', {'form': form})

            errors = []
            imported_software = []
            existing_software_names = set(Software.objects.filter(is_deleted=False).values_list('name', flat=True))

            for index, row in df.iterrows():
                row_errors = []
                try:
                    name = str(row.get('name', '')).strip()
                    category_name = str(row.get('category', '')).strip()
                    editions_names = str(row.get('editions (comma separated)', '')).strip().split(',')
                    versions_names = str(row.get('versions (comma separated)', '')).strip().split(',')
                    addons_names = str(row.get('addons (comma separated)', '')).strip().split(',')

                    if not all([name, category_name]):
                        row_errors.append("Software name and category fields are required.")

                    try:
                        category = SoftwareCategory.objects.get(name=category_name)
                    except SoftwareCategory.DoesNotExist:
                        row_errors.append(f"Category '{category_name}' does not exist.")

                    if name in existing_software_names:
                        row_errors.append(f"Software with the name '{name}' already exists.")

                    edition_instances = []
                    for edition_name in editions_names:
                        edition_name = edition_name.strip()
                        if edition_name:
                            try:
                                edition = SoftwareEdition.objects.get(name=edition_name)
                                edition_instances.append(edition)
                            except SoftwareEdition.DoesNotExist:
                                row_errors.append(f"Edition '{edition_name}' does not exist.")

                    version_instances = []
                    for version_name in versions_names:
                        version_name = version_name.strip()
                        if version_name:
                            try:
                                version = SoftwareVersion.objects.get(name=version_name)
                                version_instances.append(version)
                            except SoftwareVersion.DoesNotExist:
                                row_errors.append(f"Version '{version_name}' does not exist.")

                    addon_instances = []
                    for addon_name in addons_names:
                        addon_name = addon_name.strip()
                        if addon_name:
                            try:
                                addon = Addon.objects.get(name=addon_name)
                                addon_instances.append(addon)
                            except Addon.DoesNotExist:
                                row_errors.append(f"Addon '{addon_name}' does not exist.")

                    if row_errors:
                        errors.append(f"Row {index + 2}: {', '.join(row_errors)}")
                        continue 

                    software = Software(name=name, category=category)
                    software.save()  

                    software.editions.set(edition_instances)
                    software.versions.set(version_instances)
                    software.addons.set(addon_instances)

                    imported_software.append(software)

                except Exception as e:
                    errors.append(f"Row {index + 2}: Error importing row: {str(e)}")
                    continue

            if errors:
                messages.error(request, 'Import failed with errors. Please check the detailed error report.')
                AuditLog.objects.create(
                    user=request.user,
                    action="IMPORT FAILED",
                    details=(
                        f"Errors occurred during import on {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}: "
                        f"{', '.join(errors)}"
                    )
                )
                return render(request, 'DuxteSubscriptions/import_software.html', {'form': form, 'errors': errors})

            try:
                messages.success(request, 'Software import process completed successfully.')
                AuditLog.objects.create(
                    user=request.user,
                    action="IMPORT SUCCEEDED",
                    details="Software import process completed successfully."
                )
                return redirect('software_list')
            except Exception as e:
                messages.error(request, f"Import failed: {str(e)}")
                AuditLog.objects.create(
                    user=request.user,
                    action="IMPORT FAILED",
                    details=(
                        f"Import failed on {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}: "
                        f"{str(e)}"
                    )
                )
                return render(request, 'DuxteSubscriptions/import_software.html', {'form': form, 'errors': errors})
    else:
        form = SoftwareImportForm()

    return render(request, 'DuxteSubscriptions/import_software.html', {'form': form})

